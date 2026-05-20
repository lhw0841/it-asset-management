# monthly_report.py
from db_connect import query_to_df
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

HEADER_FILL  = PatternFill("solid", fgColor="534AB7")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
BORDER_SIDE  = Side(style='thin', color='D0CFCF')
CELL_BORDER  = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE,  bottom=BORDER_SIDE
)
ALT_FILL = PatternFill("solid", fgColor="F5F4FF")

def style_header(ws):
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = CELL_BORDER
    ws.row_dimensions[1].height = 22

def style_data(ws):
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ALT_FILL if i % 2 == 0 else None
        for cell in row:
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

def create_monthly_report():
    wb = Workbook()
    now = datetime.now()

    # Sheet 1: 전체 자산 현황
    ws1 = wb.active
    ws1.title = "자산현황"
    df1 = query_to_df("""
        SELECT asset_id AS ID, asset_name AS 장비명, type_name AS 유형,
               status AS 상태, 사용자, 부서, 사용연수, 보증상태,
               purchase_cost AS 구매가
        FROM VW_ASSET_DASHBOARD
        ORDER BY type_name, asset_name
    """)
    ws1.append(list(df1.columns))
    for row in df1.itertuples(index=False):
        ws1.append(list(row))
    style_header(ws1)
    style_data(ws1)
    auto_width(ws1)

    # Sheet 2: 미처리 장애
    ws2 = wb.create_sheet("미처리장애")
    df2 = query_to_df("""
        SELECT inc_id AS 번호, 우선순위, asset_name AS 장비,
               신고자, 부서, reported_at AS 신고일시,
               경과시간_시간 AS 경과시간, SLA상태, description AS 내용
        FROM VW_OPEN_INCIDENTS
        ORDER BY priority, reported_at
    """)
    ws2.append(list(df2.columns))
    for row in df2.itertuples(index=False):
        ws2.append(list(row))
    style_header(ws2)
    style_data(ws2)
    auto_width(ws2)

    # Sheet 3: 라이선스 만료 현황
    ws3 = wb.create_sheet("라이선스만료")
    df3 = query_to_df("""
        SELECT lic_id AS ID, sw_name AS 소프트웨어, version AS 버전,
               asset_name AS 설치장비, 사용자, 부서,
               expire_date AS 만료일, 잔여일수, 만료상태
        FROM VW_LICENSE_STATUS
        WHERE 만료상태 != '정상'
        ORDER BY expire_date
    """)
    ws3.append(list(df3.columns))
    for row in df3.itertuples(index=False):
        ws3.append(list(row))
    style_header(ws3)
    style_data(ws3)
    auto_width(ws3)

    # Sheet 4: 교체 권고 장비
    ws4 = wb.create_sheet("교체권고")
    df4 = query_to_df("""
        WITH AssetAge AS (
            SELECT a.asset_id, a.asset_name, a.purchase_date, a.purchase_cost,
                   at.type_name, at.depreciation_years,
                   d.dept_name,
                   DATEDIFF(YEAR, a.purchase_date, GETDATE()) AS age_year,
                   e.emp_name AS assigned_user
            FROM ASSET a
            JOIN ASSET_TYPE at ON a.type_id = at.type_id
            LEFT JOIN ASSET_ASSIGNMENT ag ON a.asset_id = ag.asset_id AND ag.return_date IS NULL
            LEFT JOIN EMPLOYEE e ON ag.emp_id = e.emp_id
            LEFT JOIN DEPT d ON e.dept_id = d.dept_id
            WHERE a.status != '폐기'
        ),
        ReplaceFlag AS (
            SELECT *,
                CASE
                    WHEN age_year >= depreciation_years + 2 THEN '즉시교체'
                    WHEN age_year >= depreciation_years     THEN '교체권고'
                    WHEN age_year >= depreciation_years - 1 THEN '교체예정'
                    ELSE '정상'
                END AS replace_status
            FROM AssetAge
        )
        SELECT dept_name AS 부서, type_name AS 유형, asset_name AS 장비,
               assigned_user AS 사용자, age_year AS 사용연수,
               replace_status AS 교체상태, purchase_cost AS 구매가
        FROM ReplaceFlag
        WHERE replace_status != '정상'
        ORDER BY CASE replace_status WHEN '즉시교체' THEN 1 WHEN '교체권고' THEN 2 ELSE 3 END
    """)
    ws4.append(list(df4.columns))
    for row in df4.itertuples(index=False):
        ws4.append(list(row))
    style_header(ws4)
    style_data(ws4)
    auto_width(ws4)

    # 저장
    filename = f"IT자산보고서_{now.strftime('%Y%m')}.xlsx"
    wb.save(filename)
    print(f"\n보고서 생성 완료: {filename}")
    print(f"  - 자산현황:    {len(df1)}행")
    print(f"  - 미처리장애:  {len(df2)}행")
    print(f"  - 라이선스만료: {len(df3)}행")
    print(f"  - 교체권고:    {len(df4)}행")

if __name__ == "__main__":
    create_monthly_report()